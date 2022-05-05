from openpyxl import Workbook
from openpyxl import load_workbook
import win32com.client as win32
import os

def extensao(name):
   for i in reversed(range(len(name))):
      if name[i] == '.':
         break
   return i

def conv():
   path = os.path.join(os.getcwd(),'planilhas')
   planilhas = [file for file in os.listdir(path) if file[extensao(file):] != '.xlsx']

   print(planilhas)
   cwd = os.getcwd()
   for item in planilhas:
      fname = os.path.join(cwd,'planilhas/'+item)
      excel = win32.dynamic.Dispatch('Excel.Application')
      wb = excel.Workbooks.Open(fname)
      name = 'planilhas/'+item[:extensao(item)]
      sname = os.path.join(cwd,name)
      wb.SaveAs(sname, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
      wb.Close()
      excel.Application.Quit()
      os.remove('planilhas/'+item)

if __name__ == '__main__':
   path = os.path.join(os.getcwd(), 'resultado/')
   if os.path.exists(path):
        if os.path.isdir(path):
            if os.path.isfile(os.path.join(path,'Planilhas_juntas.xlsx')):
                path = os.path.join(path,'Planilhas_juntas.xlsx')
                print("Apagando execuçao anterior")
                os.remove(path)
        else:
            os.mkdir(path)
   else:
      os.mkdir(path)
   conv()
   path = os.path.join(os.getcwd(),'planilhas')
   planilhas = [file for file in os.listdir(path)]
   wb = Workbook()
   ws = wb.active
   linha_plan = 1
   header = ['CDUNIDADEGESTORA','NMUNIDADEGESTORA','NUPREPARACAOPAGAMENTO','VLPREPARACAOPAGAMENTO','JUSTIFICATIVAQUEBRAORDEM','ORDENADORDESPESA','DTQUEBRAORDEMCRON']
   for i, h in enumerate(header):
      ws.cell(linha_plan,i+1,h)
   linha_plan = 2
   for plan in planilhas:
      print(plan)
      wbq = load_workbook('planilhas/'+plan)
      pagina = [sheet for sheet in wbq.sheetnames if sheet.find('tualizar') != -1] #buscando elemento caracteristico de todas as sheets
      try:
         wsq = wbq[pagina[0]]
      except Exception as e:
         raise ValueError('O arquivo: ' +plan+ ' não possui a planilha Atualizar portal transparencia (renomeie a correta manualmente).')
      linha = 2
      while (wsq.cell(linha,1).value != None):
         for col in range(1,8):
            if col != 5:
               ws.cell(linha_plan,col,wsq.cell(linha,col).value)
            else:
               ws.cell(linha_plan,col,'*Justi:*'+wsq.cell(linha,col).value)
         linha += 1
         linha_plan += 1
      print(linha)
   wb.save("resultado/Planilhas_juntas.xlsx")